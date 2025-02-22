
#Button 1 - Entry 5 Entry 8 Button 6
#Button 2 - Entry 5 Entry 8 Subjects Button 6
#Buton 3 - Attendance, Button 6

# 5 - 252 6 - 591 7 - 365 8 - 478
from pathlib import Path

# from tkinter import *
# Explicit imports to satisfy Flake8
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, ttk,messagebox,Image
from capture_images import ImageCapture
from architecture import *
import subprocess
import openpyxl
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0")


def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


window = Tk()

window.iconbitmap(f"C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\icon.ico")
window.title("FACE RECOGNITION ATTENDANCE TRACKER")
window.geometry("1024x768")
window.configure(bg = "#C4C4C4")

def button_1_onclick():
    for element in [button_5,button_8,button_7,subjects,attendance,entry_50,entry_80]:
        element.config(state="disabled")
        element.place_forget()
    # Check if buttons are placed and only then configure and re-place them
    for btn in [entry_5, button_6,entry_8]:
        if 'y' in btn.place_info():  # Check if 'y' exists in the place_info dictionary
            btn.config(state="normal")
            btn.place(x=708.0, y=btn.place_info()['y'], width=300.0, height=70.0)
        else:
            # Place them with initial y values if not already placed
            initial_y = 252.0 if btn == entry_5 else 365.0 if btn == entry_8 else 591.0
            btn.place(x=708.0, y=initial_y, width=300.0, height=70.0)
            btn.config(state="normal")

def button_2_onclick():
    for elements in [button_5,button_8,button_7,attendance,entry_5,entry_8]:
        elements.config(state="disabled")
        elements.place_forget()
    for btn in[entry_50,entry_80,button_6,subjects]:
        if 'y' in btn.place_info():
            btn.config(state="normal")
            btn.place(x=708.0,y=btn.place_info()['y'],width=300.0,height = 70.0)
        else:
            initial_y = 252.0 if btn == entry_50 else 365.0 if btn == entry_80 else 591.0 if btn == button_6 else 478.0 if btn ==subjects else 478.0
            btn.place(x = 708.0, y = initial_y, width= 300.0, height  = 70.0)
            btn.config(state="normal")

def button_3_onclick():
    for btn in [button_5,button_7,button_8,entry_5,entry_8,subjects,entry_80,entry_50]:
        btn.config(state="normal")
        btn.place_forget()
    for btn in [button_6,attendance]:
        if 'y' in btn.place_info():
            btn.config(state="normal")
            btn.place(x=708.0,y=btn.place_info()['y'],width=300.0,height = 70.0)
        else:
            initial_y = 591.0 if btn == button_6 else 365.0 if btn == attendance else 591.0
            btn.place(x = 708.0, y = initial_y, width= 300.0, height  = 70.0)
            btn.config(state="normal")
def button_4_onclick():
    # Disable and hide all other buttons and entries
    for btn in [button_5, button_6, button_7, button_8, entry_5, entry_8, subjects, entry_50, entry_80, attendance]:
        btn.config(state="disabled")
        btn.place_forget()

    # Define the path to the subject_attendance folder
    subject_attendance_folder = "C:\\Users\\sujun\\Documents\\Projects\\build\\subject_attendance"

    # Create the defaulters folder if it doesn't exist
    defaulters_folder = "C:\\Users\\sujun\\Documents\\Projects\\build\\defaulters"
    os.makedirs(defaulters_folder, exist_ok=True)

    # Initialize an empty DataFrame to store defaulters
    defaulter_data = {'Name': [], 'Roll No': [], 'Subject': [], 'Presence Percentage': []}

    # Iterate through each Excel file in the subject_attendance folder
    for filename in os.listdir(subject_attendance_folder):
        if filename.endswith('.xlsx'):
            excel_file_path = os.path.join(subject_attendance_folder, filename)

            # Load the Excel file
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active

            # Iterate through each row (excluding the header)
            for row in worksheet.iter_rows(min_row=2,values_only=True):
                name, roll_no, *presence_percentages = row

                # Calculate the average presence percentage if there are sessions available
                total_sessions = len(presence_percentages)
                if total_sessions > 0:
                    total_present = sum(1 for p in presence_percentages if p == 'Present')
                    presence_percentage = (total_present / total_sessions) * 100

                    # If presence percentage is less than 75%, add the defaulter to the DataFrame
                    if presence_percentage < 75:
                        subject_name = os.path.splitext(filename)[0]
                        defaulter_data['Name'].append(name)
                        defaulter_data['Roll No'].append(roll_no)
                        defaulter_data['Subject'].append(subject_name)
                        defaulter_data['Presence Percentage'].append(presence_percentage)

    # Create a DataFrame from the collected defaulter data
    defaulter_df = pd.DataFrame(defaulter_data)

    # Write the defaulter data to a new Excel file in the defaulters folder
    defaulter_filename = 'defaulters.xlsx'
    defaulter_file_path = os.path.join(defaulters_folder, defaulter_filename)
    defaulter_df.to_excel(defaulter_file_path, index=False)
    workbook = load_workbook(defaulter_file_path)
    worksheet = workbook.active

# Define color fills
    red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

# Apply color coding based on presence percentage
    for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
        presence_percentage_cell = row[3]  # Assuming the presence percentage is in the fourth column
        if presence_percentage_cell.value < 50:
            presence_percentage_cell.fill = red_fill
        elif 50 <= presence_percentage_cell.value < 75:
            presence_percentage_cell.fill = yellow_fill

# Save the changes to the workbook
    workbook.save(defaulter_file_path)
    messagebox.showinfo("Success", f"Defaulters list generated successfully. Saved as {defaulter_filename}")
    excel_filename = f"C:\\Users\\sujun\\Documents\\Projects\\build\\defaulters\\defaulters.xlsx"
    subprocess.Popen([excel_filename], shell=True)
       

canvas = Canvas(
    window,
    bg = "#C4C4C4",
    height = 768,
    width = 1024,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

def button_6_click():
    if entry_5.place_info() or entry_8.place_info():
        name = entry_5.get()
        roll_no = entry_8.get()
        if name.strip() == '' or roll_no.strip() == '':
            messagebox.showerror("Error", "Please enter name and roll number.")
            return
        try:
            int(roll_no)
        except ValueError:
            messagebox.showerror("Error", "Roll number must be a number.")
            return
        run_face_recognition(name, roll_no)
    elif attendance.place_info():
        sub = attendance.get()
        if sub.strip() == '':
            messagebox.showerror("Error", "Please Select a Subject")
        open_subject_selection_box(sub)
    elif entry_50.place_info() or entry_80.place_info():
        name = entry_50.get()
        roll_no = entry_80.get()
        subject = subjects.get()
        if name.strip() == '' or roll_no.strip() == '' or subject == '':
            messagebox.showerror("Error", "Please date,time and Select Subject")
        try:
            str(roll_no)
        except ValueError:
            messagebox.showerror("Error", "Enter valid date")
            return
        attendance_session(name,roll_no,subject)    
    return


def run_face_recognition(name, roll_no):
    import tkinter as tk
    from tkinter import messagebox
    import cv2
    import os
    import albumentations as A
    import numpy as np
    import pandas as pd

    class ImageCapture:
        def __init__(self, output_dir='Faces'):
            self.output_dir = output_dir

        def augment_image(self, image):
            transform = A.Compose([
                A.RandomBrightnessContrast(p=0.5),
                A.RandomGamma(p=0.5),
                A.OpticalDistortion(distort_limit=0.05, shift_limit=0.05, p=0.5),
            ])
            images = [transform(image=image)['image'] for _ in range(3)]
            return images

        def capture_images(self):
            script_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(script_dir, self.output_dir)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                print("Error: Unable to open the camera.")
                return

            num_images = 50
            student_info = {'Name': [], 'Roll No': []}
            images_folder = os.path.join(output_dir, name)
            if not os.path.exists(images_folder):
                os.makedirs(images_folder)

            images_taken = 0
            while images_taken < num_images:
                ret, frame = cap.read()
                if not ret:
                    print("Error: Failed to capture image.")
                    break

                cv2.imshow('Capture Image', frame)
                image_path = os.path.join(images_folder, f'{name}_{images_taken + 1}.jpg')
                cv2.imwrite(image_path, frame)
                print(f"Original Image {images_taken + 1}/{num_images} captured for {name}")
                augmented_images = self.augment_image(frame)
                for i, aug_image in enumerate(augmented_images, start=1):
                    aug_image_path = os.path.join(images_folder, f'{name}_{images_taken + 1}_aug_{i}.jpg')
                    cv2.imwrite(aug_image_path, aug_image)
                    print(f"Augmented Image {i} for {name}_{images_taken + 1} saved.")

                images_taken += 1

                if cv2.waitKey(1) == ord('q'):
                    break

            cap.release()
            cv2.destroyAllWindows()
            student_info['Name'].append(name)
            student_info['Roll No'].append(roll_no)
            df = pd.DataFrame(student_info)
            df['Roll No'] = pd.to_numeric(df['Roll No'])
            df.sort_values(by='Roll No', inplace=True)
            excel_path = 'student_info.xlsx'
            if os.path.exists(excel_path):
                existing_df = pd.read_excel(excel_path)
                df = pd.concat([existing_df, df], ignore_index=True)
                df.drop_duplicates(subset=['Roll No'], keep='first', inplace=True)
            else:
                pass

            df.sort_values(by='Roll No', inplace=True)
            df.to_excel(excel_path, index=False)
            subjects = ['CN', 'COA', 'AT', 'MATHS', 'OS']
            for subject in subjects:
                subject_excel_path = os.path.join('subject_attendance', f'{subject}_attendance.xlsx')
                if os.path.exists(subject_excel_path):
                    subject_df = pd.read_excel(subject_excel_path)
                else:
                    subject_df = pd.DataFrame(columns=['Name', 'Roll No'])
                subject_df = pd.concat([subject_df, df], ignore_index=True)
                subject_df.drop_duplicates(subset=['Roll No'], keep='first', inplace=True)
                subject_df.sort_values(by='Roll No', inplace=True)
                subject_df.to_excel(subject_excel_path, index=False)

    image_capture = ImageCapture()
    image_capture.capture_images()

def attendance_session(date,time,subject):
    import cv2
    import numpy as np
    import mtcnn
    from train_v2 import normalize, l2_normalizer
    from scipy.spatial.distance import cosine
    from tensorflow.keras.models import load_model
    import pickle
    import threading
    import pandas as pd
    from datetime import datetime, timedelta
    import os
    import openpyxl
    from openpyxl.styles import PatternFill

    class FaceRecognitionSystem:
        def __init__(self):
            self.confidence_t = 0.99
            self.recognition_t = 0.5
            self.required_size = (160, 160)
            self.face_detector = mtcnn.MTCNN()
            self.face_encoder = InceptionResNetV2()
            self.path_m = "facenet_keras_weights.h5"
            self.face_encoder.load_weights(self.path_m)
            self.encoding_dict = self.load_pickle('encodings\\encodings.pkl')

        def load_pickle(self, path):
            with open(path, 'rb') as f:
                encoding_dict = pickle.load(f)
            return encoding_dict

        def get_face(self, img, box):
            x1, y1, width, height = box
            x1, y1 = abs(x1), abs(y1)
            x2, y2 = x1 + width, y1 + height
            face = img[y1:y2, x1:x2]
            return face, (x1, y1), (x2, y2)

        def get_encode(self, face, size):
            face = normalize(face)
            face = cv2.resize(face, size)
            encode = self.face_encoder.predict(np.expand_dims(face, axis=0))[0]
            return encode

        def update_attendance(self, name, subject_attendance_file, datetime_str, buffer_lecture=False):
            lecture_date = datetime.now().strftime("%Y-%m-%d")
            unique_column_name = datetime_str
            if buffer_lecture:
                unique_column_name += "_Buffer"
            if not os.path.exists(subject_attendance_file):
                df = pd.DataFrame(columns=['Name', unique_column_name])
            else:
                df = pd.read_excel(subject_attendance_file)
                if unique_column_name not in df.columns:
                    df[unique_column_name] = None
            if name not in df['Name'].values:
                new_entry = pd.DataFrame({'Name': [name], unique_column_name: 'Present'})
                df = pd.concat([df, new_entry], ignore_index=True)
                print(f"Attendance Marked For {name}")
            else:
                df.loc[df['Name'] == name, unique_column_name] = 'Present'
                print(f"{name} Already Present.")
            df[unique_column_name] = df[unique_column_name].fillna("ABSENT")
            total_ses = len(df.columns) - 3
            df['Presence Percentage'] = df.drop(columns=['Name']).apply(lambda row: (row == 'Present').sum(), axis=1) / total_ses * 100
            cols = [col for col in df.columns if col != 'Presence Percentage']
            cols.append('Presence Percentage')
            df = df[cols]
            df.to_excel(subject_attendance_file, index=False)
            workbook = openpyxl.load_workbook(subject_attendance_file)
            worksheet = workbook.active
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            percentage_col = worksheet.max_column
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, percentage_col)
                if cell.value is not None:
                    try:
                        percentage = float(cell.value)
                        if percentage < 50:
                            cell.fill = red_fill
                        elif 50 <= percentage < 75:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = green_fill
                    except ValueError:
                        pass
            workbook.save(subject_attendance_file)

        def process_frame(self, frame, subject_name, datetime_str, buffer_lecture=False):
            img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = self.face_detector.detect_faces(img_rgb)
            subject_attendance_file = os.path.join('subject_attendance', f'{subject_name}_attendance.xlsx')
            for res in results:
                if res['confidence'] < self.confidence_t:
                    continue
                face, pt_1, pt_2 = self.get_face(img_rgb, res['box'])
                encode = self.get_encode(face, self.required_size)
                encode = l2_normalizer.transform(encode.reshape(1, -1))[0]
                name = 'unknown'
                distance = float("inf")
                for db_name, db_encode in self.encoding_dict.items():
                    dist = cosine(db_encode, encode)
                    if dist < self.recognition_t and dist < distance:
                        name = db_name
                        distance = dist
                        self.update_attendance(name, subject_attendance_file, datetime_str, buffer_lecture=buffer_lecture)
                if name == 'unknown':
                    cv2.rectangle(frame, pt_1, pt_2, (0, 0, 255), 2)
                    cv2.putText(frame, name, pt_1, cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 1)
                else:
                    cv2.rectangle(frame, pt_1, pt_2, (0, 255, 0), 2)
                    cv2.putText(frame, name + f'__{distance:.2f}', (pt_1[0], pt_1[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 1,
                                (0, 200, 200), 2)
            return frame

        def camera_thread(self, subject_name, datetime_str):
            cap = cv2.VideoCapture(0)
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    messagebox.showerror("Error","CAMERA NOT OPENED")
                    break
                current_time = datetime.now().time()
                buffer_lecture = False
                if current_time >= datetime.strptime("17:00:00", "%H:%M:%S").time():
                    buffer_lecture = True
                frame = self.process_frame(frame, subject_name, datetime_str, buffer_lecture)
                cv2.imshow('camera', frame)
                if cv2.waitKey(1) & 0xFF in [ord('q'), ord('Q')]:
                    break
            cap.release()
            cv2.destroyAllWindows()

    if __name__ == "__main__":
        os.makedirs('subject_attendance', exist_ok=True)
        frc = FaceRecognitionSystem()
        subject_name = subject
        datetime_str = date
        if not subject_name:
            messagebox.showerror("Error","Please Enter the Subject Name")
        else:
            frc.camera_thread(subject_name, datetime_str)

def open_subject_selection_box(subject):
            excel_filename = f"C:\\Users\\sujun\\Documents\\Projects\\build\\subject_attendance\\{subject}_attendance.xlsx"
            try:
                subprocess.Popen([excel_filename], shell=True)
                # workbook = load_workbook(excel_filename)
                # Add code to handle the workbook, e.g., display or process its contents
                messagebox.showinfo("Success", f"Opened {subject} attendance Excel file")
            except FileNotFoundError:
                messagebox.showerror("Error", f"Excel file for {subject} not found")
            

    
    

canvas.place(x = 0, y = 0)
image_image_1 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\image_1.png"))
image_1 = canvas.create_image(
    512.0,
    422.0,
    image=image_image_1
)

button_image_1 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    text="REGISTER STUDENT",
    command= button_1_onclick,
    relief="flat",
    compound="center"
)
button_1.place(
    x=69.0,
    y=252.0,
    width=300.0,
    height=70.0
)

button_image_2 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    text="ATTENDANCE SESSION",
    command=button_2_onclick,
    relief="flat",
    compound="center"
)
button_2.place(
    x=69.0,
    y=365.0,
    width=300.0,
    height=70.0
)

button_image_3 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    text="ATTENDANCE VIEWER",
    command=button_3_onclick,
    relief="flat",
    compound="center"
)
button_3.place(
    x=69.0,
    y=478.0,
    width=300.0,
    height=70.0
)

button_image_4 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    text="DEFAULTER GENERATOR",
    font=("Roboto", 14 * -1),
    command=button_4_onclick,
    relief="flat",
    compound="center"
)
button_4.place(
    x=69.0,
    y=591.0,
    width=300.0,
    height=70.0
)

canvas.create_rectangle(
    0.0,
    0.0,
    1024.0,
    78.0,
    fill="#151019",
    outline="")

canvas.create_text(
    7.0,
    18.0,
    anchor="nw",
    text="FACE RECOGNITION ATTENDANCE SYSTEM",
    fill="#FFFFFF",
    font=("Inter BlackItalic", 32 * -1)
)
image_logo = PhotoImage(file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\sies_logo.png"))
canvas.create_image(
    (891.0+1024.0)/2,
    (0.0+109.0)/2,
    image = image_logo
)

button_image_5 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_5.png"))
button_5 = Button(
    image=button_image_5,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: print("button_5 clicked"),
    relief="flat",
    state="disabled",
)
button_5.place(
    x=708.0,
    y=252.0,
    width=300.0,
    height=70.0,
)
button_5.place_forget()

button_image_6 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_6.png"))
button_6 = Button(
    image=button_image_6,
    borderwidth=0,
    highlightthickness=0,
    command=button_6_click,
    relief="flat",
    state="disabled",
    text="SUBMIT",
    font=("Roboto", 15 * -1),
    compound="center"
)
button_6.place(
    x=708.0,
    y=591.0,
    width=300.0,
    height=70.0
)
button_6.place_forget()

button_image_7 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_7.png"))
button_7 = Button(
    image=button_image_7,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: print("button_7 clicked"),
    relief="flat",
    state="disabled"
)
button_7.place(
    x=708.0,
    y=478.0,
    width=300.0,
    height=70.0
)
button_7.place_forget()

button_image_8 = PhotoImage(
    file=relative_to_assets("C:\\Users\\sujun\\Documents\\Projects\\build\\assets\\frame0\\button_8.png"))
button_8 = Button(
    image=button_image_8,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: print("button_8 clicked"),
    relief="flat",
    state= "disabled"
)
button_8.place(
    x=708.0,
    y=365.0,
    width=300.0,
    height=70.0
)
button_8.place_forget()

entry_5 = Entry(window,bd=0,state="normal",font=("Arial", 14, "bold"))
entry_5.insert(0,"ENTER NAME")
entry_5.bind("<FocusIn>",lambda event,entry=entry_5,label="ENTER NAME":on_entry_click(event,entry,label))
entry_5.bind("<FocusOut>", lambda event, entry=entry_5, label="ENTER NAME": on_focusout(event, entry,label))

entry_5.place(x=708.0,y=252.0,width=300.0,height=70.0)
entry_5.place_forget()

entry_8 = Entry(window,bd=0,state="normal",font=("Arial", 14, "bold"))
entry_8.insert(0,"ENTER ROLL NO")
entry_8.bind("<FocusIn>",lambda event,entry = entry_8,label = "ENTER ROLL NO":on_entry_click(event,entry,label))
entry_8.bind("<FocusOut>",lambda event,entry = entry_8,label = "ENTER ROLL NO":on_focusout(event,entry,label))
entry_8.place(x=708.0,y=365.0,width=300.0,height=70.0)
entry_8.place_forget()

subjects_list = ["MATHS","COA","OS","AT","CN"]
subjects = ttk.Combobox(window,values=subjects_list,state="disabled",font=("Arial", 14, "bold"))
subjects.place(x=708.0,
    y=478.0,
    width=300.0,
    height=70.0)
subjects.place_forget()

attendance = ttk.Combobox(window,values=subjects_list,state="disabled",font=("Arial", 14, "bold"))
attendance.place(x = 708.0, y = 365.0, width = 300.0, height = 70.0)
attendance.place_forget()

entry_50 = Entry(window,bd=0,state="normal",font=("Arial", 14, "bold"))
entry_50.place(x=708.0,y=252.0,width=300.0,height=70.0)
entry_50.insert(0,"ENTER DATE IN YYYY:MM:DD")
entry_50.bind("<FocusIn>",lambda event,entry = entry_50,label = "ENTER DATE IN YYYY:MM:DD":on_entry_click(event,entry,label))
entry_50.bind("<FocusOut>",lambda event,entry = entry_50,label = "ENTER DATE IN YYYY:MM:DD":on_focusout(event,entry,label))
entry_50.place_forget()

entry_80 = Entry(window,bd=0,state="normal",font=("Arial", 14, "bold"))
entry_80.place(x=708.0,y=365.0,width=300.0,height=70.0)
entry_80.insert(0,"ENTER TIME HH:MM")
entry_80.bind("<FocusIn>",lambda event,entry = entry_80,label = "ENTER TIME HH:MM":on_entry_click(event,entry,label))
entry_80.bind("<FocusOut>",lambda event,entry = entry_80,label = "ENTER TIME HH:MM":on_focusout(event,entry,label))
entry_80.place_forget()

def on_entry_click(event, entry, label):
        if entry.get() == label:
            entry.delete(0, "end")
            entry.config(fg="black")

def on_focusout(event, entry, label):
        if entry.get() == "":
            entry.insert(0, label)
            entry.config(fg="gray")

window.resizable(False, False)
window.mainloop()
